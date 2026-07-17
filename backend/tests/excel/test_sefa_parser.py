"""
Testes unitários para o parser da planilha SEFA-PA.

Formato real (conforme análise dos arquivos modelo SEFA-PA):
  - Tipo de antecipação vem do TÍTULO (linha 0), não de uma coluna
  - Múltiplas linhas por NF (um item por produto) → agregação por DANFE
  - Coluna de valor: "TOTAL ESTORNO" (Formato A) ou "ICMS A PAGAR" (Formato B)
  - Suporte a .xls e .xlsx
"""
from __future__ import annotations

from decimal import Decimal
from pathlib import Path

import openpyxl
import pytest

from app.excel.sefa_parser import (
    _extract_tipo_from_title,
    _map_columns,
    parse_sefa_excel,
)


CHAVE_NF_A = "35260304165376000107550010001554691944403164"
CHAVE_NF_B = "11250903680934000100550010000271461900121158"
CNPJ_A = "04165376000107"


def _make_xlsx_formato_a(tmp_path: Path, dados: list[list], titulo: str | None = None) -> Path:
    """Cria planilha no Formato A (produto a produto, coluna TOTAL ESTORNO)."""
    if titulo is None:
        titulo = "Receita: COD1173 ICMS ANTECIPADO ESPECIAL"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append([None, titulo])   # linha 0: título
    ws.append([              # linha 1: cabeçalho
        None, "Nº ITEM", "DESCRIÇÃO DO PRODUTO", "NF", "CNPJ ORIGEM",
        "NCM", "PRODUTO", "VALOR NOTA FISCAL", "VALOR DO PRODUTO",
        "VALOR BASE CALC", "VALOR ICMS CALC", "ICMS DO PRODUTO",
        "EXPECTATIVA A PAGAR", "CODIGO DE APURAÇÃO", "ALIQUOTA",
        "BASE APURADA", "DANFE", "BASE ESTORNO", "ALIQUOTA", "TOTAL ESTORNO",
    ])
    for row in dados:
        ws.append(row)
    p = tmp_path / "sefa_a.xlsx"
    wb.save(str(p))
    return p


def _make_xlsx_formato_b(tmp_path: Path, dados: list[list], titulo: str | None = None) -> Path:
    """Cria planilha no Formato B (nota a nota, coluna ICMS A PAGAR)."""
    if titulo is None:
        titulo = "Receita: 1173 - Antecipado Especial"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append([None, titulo])
    ws.append([
        None, "Nº ITEM", "SERIE / NOTA", "NCM", "CNPJ",
        "VALOR NOTA FISCAL", "VALOR NOTA PRODUTO", "ICMS DO PRODUTO",
        "VALOR BASE CALC", "VALOR ICMS CALC", "ICMS A PAGAR", "BC Portaria", "DANFE",
    ])
    for row in dados:
        ws.append(row)
    p = tmp_path / "sefa_b.xlsx"
    wb.save(str(p))
    return p


# ── Extração de tipo do título ────────────────────────────────────────────────

class TestExtractTipo:
    def test_cod1173_especial(self) -> None:
        assert _extract_tipo_from_title(["", "Receita: COD1173 ICMS ANTECIPADO ESPECIAL"]) == "ESPECIAL"

    def test_1173_especial(self) -> None:
        assert _extract_tipo_from_title(["", "Receita: 1173 - Antecipado Especial"]) == "ESPECIAL"

    def test_1146_normal(self) -> None:
        assert _extract_tipo_from_title(["", "Receita: 1146 ICMS Antecipado Normal"]) == "NORMAL"

    def test_cod1152_cesta(self) -> None:
        assert _extract_tipo_from_title(["", "COD1152 Cesta Basica"]) == "CESTA_BASICA"

    def test_raises_unknown(self) -> None:
        with pytest.raises(ValueError, match="Tipo de antecipação não identificado"):
            _extract_tipo_from_title(["", "Sem informação de receita"])


# ── Mapeamento de colunas com prioridade ─────────────────────────────────────

class TestColumnMapping:
    def test_total_estorno_wins_over_valor_icms_calc(self) -> None:
        header = [None, "NF", "CNPJ", "VALOR ICMS CALC", "DANFE", "TOTAL ESTORNO"]
        cm = _map_columns(header)
        # TOTAL ESTORNO (prioridade 10) deve vencer VALOR ICMS CALC (prioridade 3)
        valor_col = [k for k, v in cm.items() if v == "valor_icms"]
        assert len(valor_col) == 1
        assert header[valor_col[0]] == "TOTAL ESTORNO"

    def test_icms_a_pagar_wins_over_expectativa(self) -> None:
        header = [None, "EXPECTATIVA A PAGAR", "ICMS A PAGAR", "DANFE"]
        cm = _map_columns(header)
        valor_col = [k for k, v in cm.items() if v == "valor_icms"]
        assert header[valor_col[0]] == "ICMS A PAGAR"

    def test_danfe_detected(self) -> None:
        header = [None, "NF", "CNPJ ORIGEM", "DANFE", "TOTAL ESTORNO"]
        cm = _map_columns(header)
        chave_col = [k for k, v in cm.items() if v == "chave_nfe"]
        assert len(chave_col) == 1
        assert header[chave_col[0]] == "DANFE"


# ── Formato A: produto a produto ─────────────────────────────────────────────

def _row_a(nf: str, cnpj: str, danfe: str, total_estorno: float, item: int = 1) -> list:
    return [
        None, str(item), "PRODUTO TESTE", nf, cnpj,
        "12345678", "PRODUTO", "1000", "500", "0", "0", "0", "0",
        "COD1173", "0.07", "350", danfe, "500", "0.07", total_estorno,
    ]


def test_formato_a_single_nf(tmp_path: Path) -> None:
    p = _make_xlsx_formato_a(tmp_path, [
        _row_a("000155469", "04.165.376/0001-07", CHAVE_NF_A, 124.95, 1),
        _row_a("000155469", "04.165.376/0001-07", CHAVE_NF_A, 141.12, 2),
        _row_a("000155469", "04.165.376/0001-07", CHAVE_NF_A, 52.50, 3),
    ])
    results = parse_sefa_excel(p)

    assert len(results) == 1
    r = results[0]
    assert r.chave_nfe == CHAVE_NF_A
    assert r.tipo == "ESPECIAL"
    # Soma os três itens
    from decimal import Decimal
    assert r.valor_icms == Decimal("124.95") + Decimal("141.12") + Decimal("52.50")


def test_formato_a_multiple_nfs(tmp_path: Path) -> None:
    chave2 = "31260417359233000188550010247530001340515105"
    p = _make_xlsx_formato_a(tmp_path, [
        _row_a("000155469", "04.165.376/0001-07", CHAVE_NF_A, 337.47, 1),
        _row_a("024753000", "17.359.233/0001-88", chave2, 154.94, 1),
    ])
    results = parse_sefa_excel(p)
    assert len(results) == 2
    chaves = {r.chave_nfe for r in results}
    assert CHAVE_NF_A in chaves
    assert chave2 in chaves


def test_formato_a_cnpj_cleaned(tmp_path: Path) -> None:
    p = _make_xlsx_formato_a(tmp_path, [
        _row_a("000155469", "04.165.376/0001-07", CHAVE_NF_A, 100.00, 1),
    ])
    results = parse_sefa_excel(p)
    assert results[0].emitente_cnpj == CNPJ_A


def test_formato_a_tipo_normal(tmp_path: Path) -> None:
    p = _make_xlsx_formato_a(tmp_path, [
        _row_a("001234567", "04.165.376/0001-07", CHAVE_NF_A, 200.00),
    ], titulo="Receita: 1146 ICMS Antecipado Normal")
    results = parse_sefa_excel(p)
    assert results[0].tipo == "NORMAL"


def test_formato_a_ignores_zero_total_estorno(tmp_path: Path) -> None:
    p = _make_xlsx_formato_a(tmp_path, [
        _row_a("000155469", "04.165.376/0001-07", CHAVE_NF_A, 0.0, 1),
    ])
    with pytest.raises(ValueError, match="Nenhuma antecipação"):
        parse_sefa_excel(p)


# ── Formato B: nota a nota ───────────────────────────────────────────────────

def _row_b(serie_nota: str, cnpj: str, danfe: str, icms_a_pagar: float, item: int = 1) -> list:
    return [
        None, str(item), serie_nota, "87163900", cnpj,
        "14630", "14630", "1755.6", "14630", "2779.7", icms_a_pagar, "", danfe,
    ]


def test_formato_b_single_nf(tmp_path: Path) -> None:
    p = _make_xlsx_formato_b(tmp_path, [
        _row_b("1/27146", "03.680.934/0001-00", CHAVE_NF_B, 1024.10),
    ])
    results = parse_sefa_excel(p)
    assert len(results) == 1
    r = results[0]
    assert r.numero_nf == "27146"
    assert r.serie == "1"
    assert r.chave_nfe == CHAVE_NF_B
    assert r.tipo == "ESPECIAL"
    assert r.valor_icms == Decimal("1024.1")


def test_formato_b_zero_icms_excluded(tmp_path: Path) -> None:
    """Itens com ICMS A PAGAR = 0 são excluídos do total."""
    p = _make_xlsx_formato_b(tmp_path, [
        _row_b("1/27146", "03.680.934/0001-00", CHAVE_NF_B, 1024.10, 1),
        _row_b("1/1506005", "07.069.487/0003-70", "21250907069487000370550010015060051138128114", 0.0, 1),
    ])
    results = parse_sefa_excel(p)
    # Apenas NF 27146 tem valor > 0
    assert len(results) == 1
    assert results[0].numero_nf == "27146"


def test_formato_b_serie_nota_parsed(tmp_path: Path) -> None:
    p = _make_xlsx_formato_b(tmp_path, [
        _row_b("3/98765", "03.680.934/0001-00", CHAVE_NF_B, 500.00),
    ])
    results = parse_sefa_excel(p)
    assert results[0].serie == "3"
    assert results[0].numero_nf == "98765"


# ── Erros esperados ──────────────────────────────────────────────────────────

def test_raises_if_no_danfe_column(tmp_path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append([None, "Receita: COD1173"])
    ws.append(["NF", "CNPJ", "VALOR"])
    ws.append(["001", "04165376000107", "100"])
    p = tmp_path / "sem_danfe.xlsx"
    wb.save(str(p))
    with pytest.raises(ValueError, match="DANFE"):
        parse_sefa_excel(p)


def test_raises_if_unknown_tipo(tmp_path: Path) -> None:
    p = _make_xlsx_formato_a(tmp_path, [
        _row_a("000155469", "04.165.376/0001-07", CHAVE_NF_A, 100.00),
    ], titulo="Receita: SEM CODIGO")
    with pytest.raises(ValueError, match="Tipo de antecipação não identificado"):
        parse_sefa_excel(p)


def test_dare_fields_empty(tmp_path: Path) -> None:
    """Planilha SEFA-PA não inclui DAR/DARE — campos devem ser None."""
    p = _make_xlsx_formato_a(tmp_path, [
        _row_a("000155469", "04.165.376/0001-07", CHAVE_NF_A, 100.00),
    ])
    results = parse_sefa_excel(p)
    assert results[0].dare_numero is None
    assert results[0].dare_vencimento is None
